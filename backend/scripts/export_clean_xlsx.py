"""Xuất bản xlsx sạch + enriched: backend/data/Spec_cate_gia.cleaned.xlsx.

File gốc data/Spec_cate_gia.xlsx không bị đụng tới. Mỗi sheet giữ nguyên tên và
thứ tự cột gốc (đã làm sạch: placeholder -> ô trống, cột rác bị loại, giá trị
ngoài cửa sổ hợp lý -> trống + cột "... (nguyên văn)"), sau đó là các cột dẫn
xuất từ làm sạch, cuối cùng là khối cột enrichment "(crawl)".
"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import pandas as pd
from openpyxl import Workbook

from app.catalog.clean_rules import clean_sheet
from app.catalog.enrich import load_crawl_records
from app.agent_core.search_description import (
    DESCRIPTION_COLUMN,
    build_search_description,
    select_description_fields,
)
from app.config import get_settings

_CRAWL_COLS = [
    "tên sản phẩm", "giá hiệu lực", "nguồn giá",
    "rating", "lượt bán", "bảo hành", "khuyến mãi", "url", "ảnh",
]


def _key(value) -> str | None:
    text = str(value).strip() if value is not None else ""
    if text.endswith(".0"):
        text = text[:-2]
    if not text or text.lower() in {"nan", "none", "null"} or set(text) == {"9"}:
        return None
    return text


def _crawl_values(row: dict, crawl: dict | None) -> dict:
    original = crawl.get("original_price") if crawl else None
    sale = crawl.get("sale_price") if crawl else None
    sale_valid = sale and (not original or sale <= original)
    effective = sale if sale_valid else original
    source = None
    if effective:
        source = "crawl dienmayxanh"
        as_of = str(crawl.get("crawled_at") or "")[:10]
        if as_of:
            source += f", ngày {as_of}"
    else:
        effective = row.get("giá khuyến mãi") or row.get("giá gốc")
        if effective:
            source = "thông số nhà sản xuất"
    return {
        "tên sản phẩm": crawl.get("name") if crawl else None,
        "giá hiệu lực": effective,
        "nguồn giá": source,
        "rating": crawl.get("rating") if crawl else None,
        "lượt bán": crawl.get("quantity_sold") if crawl else None,
        "bảo hành": crawl.get("warranty_policy") if crawl else None,
        "khuyến mãi": crawl.get("promotion") if crawl else None,
        "url": crawl.get("url") if crawl else None,
        "ảnh": crawl.get("image_url") if crawl else None,
    }


def main():
    s = get_settings()
    crawl_records = load_crawl_records(s.crawl_cleaned_path, s.crawl_path)
    crawl_by_pid = {_key(row.get("product_id")): row for row in crawl_records
                    if _key(row.get("product_id"))}
    crawl_by_code = {_key(row.get("product_code")): row for row in crawl_records
                     if _key(row.get("product_code"))}

    wb = Workbook()
    wb.remove(wb.active)
    xls = pd.ExcelFile(s.spec_source_path)
    total = 0
    for sheet_name in xls.sheet_names:
        source_rows = pd.read_excel(xls, sheet_name=sheet_name).to_dict(orient="records")
        rows, _ = clean_sheet(sheet_name, source_rows)
        ws = wb.create_sheet(title=sheet_name)
        # thứ tự cột: theo dòng gốc, cột dẫn xuất mới gặp nối vào sau
        headers: list[str] = []
        for row in rows:
            for k in row:
                if k not in headers and k != "display_name_tong_hop":
                    headers.append(k)
        description_fields = select_description_fields(rows)
        headers += [DESCRIPTION_COLUMN, *_CRAWL_COLS]
        ws.append(headers)
        for row in rows:
            pid = _key(row.get("productidweb"))
            sku = _key(row.get("sku"))
            crawl = (crawl_by_pid.get(pid) if pid else None) or crawl_by_code.get(sku)
            extra = _crawl_values(row, crawl)
            extra[DESCRIPTION_COLUMN] = build_search_description(
                sheet_name, str(row.get("brand") or row.get("brand_id") or ""),
                row, description_fields
            )
            ws.append([row.get(h) if h not in extra else extra[h] for h in headers])
        total += len(rows)

    out = os.path.join(os.path.dirname(s.catalog_path), "Spec_cate_gia.cleaned.xlsx")
    wb.save(out)
    print(f"Da xuat {total} dong / {len(wb.sheetnames)} sheet -> {out}")


if __name__ == "__main__":
    main()
